#!/usr/bin/env python3
"""
Evaluation script for SpreadsheetBench outputs.

This script evaluates agent outputs against SpreadsheetBench ground truth,
adapting to the output structure used by run_spreadsheetbench.py.

Usage:
    python evaluate_outputs.py --data_path data/sample_data_200 --output_dir outputs/spreadsheetbench
    
    # With optional Excel recalculation (Windows only)
    python evaluate_outputs.py --data_path data/sample_data_200 --output_dir outputs/spreadsheetbench --open_excel
"""

import argparse
import datetime
import json
import os
import sys
from collections import defaultdict

import openpyxl
from tqdm import tqdm


# ============================================================================
# Cell comparison utilities (adapted from SpreadsheetBench evaluation.py)
# ============================================================================

def datetime_to_float(dt):
    """Convert datetime to Excel float representation."""
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    """Normalize cell value for comparison."""
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_cell_value(v1, v2):
    """Compare two cell values after normalization."""
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        return False
    return v1 == v2


def col_num2name(n):
    """Convert a column number to an Excel column name."""
    name = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    """Convert an Excel column name to a column number."""
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num


def parse_cell_range(range_str):
    """Parse a range string like 'A1:AB12'."""
    start_cell, end_cell = range_str.split(':')
    start_col, start_row = '', ''
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char

    end_col, end_row = '', ''
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return start_col, start_row, end_col, end_row


def generate_cell_names(range_str, max_row=None, max_col=None):
    """Generate a list of all cell names in the specified range."""
    if ':' not in range_str:
        return [range_str]

    start_col, start_row, end_col, end_row = parse_cell_range(range_str)

    # Handle full column range (e.g., "A:B" - no row numbers)
    if not start_row and not end_row:
        start_row = '1'
        end_row = str(max_row) if max_row else '1000'  # Reasonable default
    elif not start_row:
        start_row = '1'
    elif not end_row:
        end_row = str(max_row) if max_row else '1000'

    # Handle full row range (e.g., "1:10" - no column letters)
    if not start_col and not end_col:
        start_col = 'A'
        end_col = col_num2name(max_col) if max_col else 'Z'  # Reasonable default
    elif not start_col:
        start_col = 'A'
    elif not end_col:
        end_col = col_num2name(max_col) if max_col else 'Z'

    start_col_num = col_name2num(start_col)
    end_col_num = col_name2num(end_col)
    start_row_num = int(start_row)
    end_row_num = int(end_row)

    columns = [col_num2name(i) for i in range(start_col_num, end_col_num + 1)]
    cell_names = [f"{col}{row}" for col in columns for row in range(start_row_num, end_row_num + 1)]
    return cell_names


def cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range):
    """Compare cells in a specific range between two workbooks."""
    if sheet_name not in wb_proc.sheetnames:
        return False, f"Worksheet '{sheet_name}' not found in output"
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]

    # Use ground truth worksheet dimensions for full column/row ranges
    cell_names = generate_cell_names(cell_range, max_row=ws_gt.max_row, max_col=ws_gt.max_column)

    for cell_name in cell_names:
        cell_gt = ws_gt[cell_name]
        cell_proc = ws_proc[cell_name]

        if not compare_cell_value(cell_gt.value, cell_proc.value):
            msg = f"Value mismatch at {cell_name}: expected '{cell_gt.value}', got '{cell_proc.value}'"
            return False, msg

    return True, ""


def compare_workbooks(gt_file, output_file, answer_position):
    """Compare output workbook against ground truth."""
    if not os.path.exists(output_file):
        return False, "Output file not found"
    
    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=output_file, data_only=True)
    except Exception as e:
        return False, f"Error loading workbook: {e}"

    # Parse answer position (may contain multiple ranges separated by comma)
    sheet_cell_ranges = answer_position.split(',')
    
    for sheet_cell_range in sheet_cell_ranges:
        sheet_cell_range = sheet_cell_range.strip()
        if '!' in sheet_cell_range:
            sheet_name, cell_range = sheet_cell_range.split('!')
            sheet_name = sheet_name.strip("'")
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = sheet_cell_range
        
        cell_range = cell_range.strip("'")
        
        result, msg = cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range)
        if not result:
            return False, msg

    return True, ""


# ============================================================================
# Excel file opening (Windows only, for formula recalculation)
# ============================================================================

def open_excel_file(filepath):
    """Open and save Excel file to trigger formula recalculation (Windows only)."""
    try:
        from win32com.client import Dispatch
    except ImportError:
        print("Warning: win32com not available. Skipping Excel recalculation.")
        return False
    
    filepath = os.path.abspath(filepath)
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlApp.DisplayAlerts = False
    xlApp.ScreenUpdating = False
    try:
        xlBook = xlApp.Workbooks.Open(Filename=filepath, UpdateLinks=False, ReadOnly=False)
        xlBook.Save()
        xlBook.Close(SaveChanges=True)
        return True
    except Exception as e:
        print(f"Error opening {filepath}: {e}")
        return False
    finally:
        xlApp.Quit()


def open_all_outputs(output_dir):
    """Open all Excel files in output directory to recalculate formulas."""
    print("Opening Excel files to recalculate formulas...")
    count = 0
    for root, dirs, files in os.walk(output_dir):
        for f in files:
            if f.endswith('.xlsx'):
                filepath = os.path.join(root, f)
                if open_excel_file(filepath):
                    count += 1
    print(f"Processed {count} Excel files")


# ============================================================================
# Data loading
# ============================================================================

def load_dataset(data_path):
    """Load dataset from JSON or JSONL file."""
    # Find data file
    candidates = [
        os.path.join(data_path, "dataset.json"),
        os.path.join(data_path, "data.json"),
        os.path.join(data_path, "data.jsonl"),
        os.path.join(data_path, "sample_data.jsonl"),
    ]
    
    if data_path.endswith(".json") or data_path.endswith(".jsonl"):
        candidates.insert(0, data_path)
    
    data_file = None
    for path in candidates:
        if os.path.exists(path):
            data_file = path
            break
    
    if data_file is None:
        # Search directory for any json/jsonl file
        if os.path.isdir(data_path):
            for f in os.listdir(data_path):
                if f.endswith(".json") or f.endswith(".jsonl"):
                    data_file = os.path.join(data_path, f)
                    break
    
    if data_file is None:
        raise FileNotFoundError(f"No dataset file found in {data_path}")
    
    print(f"Loading dataset from: {data_file}")
    
    with open(data_file, "r", encoding="utf-8") as f:
        if data_file.endswith(".json"):
            return json.load(f)
        else:
            return [json.loads(line) for line in f if line.strip()]


def find_spreadsheet_dir(data_path, instance):
    """Find the spreadsheet directory for an instance."""
    data_dir = data_path if os.path.isdir(data_path) else os.path.dirname(data_path)
    instance_id = str(instance["id"])
    spreadsheet_path = str(instance.get("spreadsheet_path", instance_id))
    
    candidates = [
        os.path.join(data_dir, "spreadsheet", spreadsheet_path),
        os.path.join(data_dir, spreadsheet_path),
        os.path.join(data_dir, "spreadsheet", instance_id),
        os.path.join(data_dir, instance_id),
    ]
    
    for path in candidates:
        if os.path.exists(path):
            return path
    
    return None


# ============================================================================
# Main evaluation
# ============================================================================

def evaluate(data_path, output_dir, start_idx=0, end_idx=None, verbose=False):
    """
    Evaluate outputs against ground truth.
    
    Returns:
        dict with evaluation results
    """
    dataset = load_dataset(data_path)
    
    if end_idx is None:
        end_idx = len(dataset)
    dataset = dataset[start_idx:end_idx]
    
    print(f"Evaluating {len(dataset)} instances...")
    
    results = []
    total_test_cases = 0
    passed_test_cases = 0
    fully_correct = 0
    
    for instance in tqdm(dataset):
        instance_id = instance["id"]
        spreadsheet_path = instance.get("spreadsheet_path", instance_id)
        answer_position = instance.get("answer_position", "")
        
        if not answer_position:
            if verbose:
                print(f"Warning: No answer_position for {instance_id}, skipping")
            continue
        
        # Find spreadsheet directory (contains ground truth)
        spreadsheet_dir = find_spreadsheet_dir(data_path, instance)
        if spreadsheet_dir is None:
            results.append({
                "id": instance_id,
                "success": False,
                "error": "Spreadsheet directory not found",
                "test_cases": [],
            })
            continue
        
        # Find output directory for this instance
        # spreadsheet_path may include a prefix like "spreadsheet/13-1",
        # so try multiple candidates to avoid double-nesting
        output_candidates = [
            os.path.join(output_dir, spreadsheet_path),
            os.path.join(output_dir, instance_id),
        ]
        output_instance_dir = output_candidates[0]
        for candidate in output_candidates:
            if os.path.isdir(candidate):
                output_instance_dir = candidate
                break
        
        # Find all test cases (ground truth files)
        # Standard format: *_answer.xlsx, Verified format: *_golden.xlsx
        all_files = os.listdir(spreadsheet_dir)
        gt_files = sorted([
            f for f in all_files
            if f.endswith("_answer.xlsx")
        ])

        if not gt_files:
            # Try verified dataset format
            gt_files = sorted([
                f for f in all_files
                if f.endswith("_golden.xlsx")
            ])

        if not gt_files:
            # Try exact match for simple naming: golden.xlsx
            if "golden.xlsx" in all_files:
                gt_files = ["golden.xlsx"]

        if not gt_files:
            results.append({
                "id": instance_id,
                "success": False,
                "error": "No ground truth files found (expected *_answer.xlsx or *_golden.xlsx)",
                "test_cases": [],
            })
            continue
        
        test_case_results = []
        
        for gt_file in gt_files:
            # Derive output filename from ground truth filename
            # Standard: "1_Q001_answer.xlsx" -> "1_Q001_output.xlsx"
            # Verified: "1_13-1_golden.xlsx" -> "1_13-1_output.xlsx"
            # Simple: "golden.xlsx" -> "initial_output.xlsx"
            if gt_file.endswith("_answer.xlsx"):
                output_file = gt_file.replace("_answer.xlsx", "_output.xlsx")
            elif gt_file == "golden.xlsx":
                output_file = "initial_output.xlsx"
            else:  # _golden.xlsx
                output_file = gt_file.replace("_golden.xlsx", "_output.xlsx")
            
            gt_path = os.path.join(spreadsheet_dir, gt_file)
            output_path = os.path.join(output_instance_dir, output_file)
            
            total_test_cases += 1
            
            result, msg = compare_workbooks(gt_path, output_path, answer_position)
            
            test_case_results.append({
                "gt_file": gt_file,
                "output_file": output_file,
                "passed": result,
                "message": msg,
            })
            
            if result:
                passed_test_cases += 1
            elif verbose:
                print(f"  {instance_id}/{output_file}: {msg}")
        
        # Calculate metrics for this instance
        passed_count = sum(1 for tc in test_case_results if tc["passed"])
        total_count = len(test_case_results)
        soft_score = passed_count / total_count if total_count > 0 else 0
        hard_score = 1 if passed_count == total_count else 0
        
        if hard_score == 1:
            fully_correct += 1
        
        results.append({
            "id": instance_id,
            "success": hard_score == 1,
            "test_cases": test_case_results,
            "passed_count": passed_count,
            "total_count": total_count,
            "soft_score": soft_score,
            "hard_score": hard_score,
        })
    
    # Calculate overall metrics
    total_instances = len(results)
    
    soft_scores = [r.get("soft_score", 0) for r in results if "soft_score" in r]
    hard_scores = [r.get("hard_score", 0) for r in results if "hard_score" in r]
    
    avg_soft_score = sum(soft_scores) / len(soft_scores) if soft_scores else 0
    avg_hard_score = sum(hard_scores) / len(hard_scores) if hard_scores else 0
    
    summary = {
        "total_instances": total_instances,
        "fully_correct_instances": fully_correct,
        "instance_accuracy": fully_correct / total_instances if total_instances > 0 else 0,
        "total_test_cases": total_test_cases,
        "passed_test_cases": passed_test_cases,
        "test_case_accuracy": passed_test_cases / total_test_cases if total_test_cases > 0 else 0,
        "avg_soft_score": avg_soft_score,
        "avg_hard_score": avg_hard_score,
    }
    
    return {
        "summary": summary,
        "results": results,
    }


def main():
    parser = argparse.ArgumentParser(description="Evaluate SpreadsheetBench outputs")
    parser.add_argument(
        "--data_path",
        type=str,
        required=True,
        help="Path to SpreadsheetBench data directory",
    )
    parser.add_argument(
        "--output_dir",
        type=str,
        required=True,
        help="Directory containing agent outputs",
    )
    parser.add_argument(
        "--results_file",
        type=str,
        default=None,
        help="Path to save evaluation results JSON (default: output_dir/eval_results.json)",
    )
    parser.add_argument(
        "--start_idx",
        type=int,
        default=0,
        help="Start index for evaluation",
    )
    parser.add_argument(
        "--end_idx",
        type=int,
        default=None,
        help="End index for evaluation (exclusive)",
    )
    parser.add_argument(
        "--open_excel",
        action="store_true",
        help="Open Excel files to recalculate formulas (Windows only)",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print detailed error messages",
    )
    args = parser.parse_args()
    
    # Optionally open Excel files to recalculate formulas
    if args.open_excel:
        open_all_outputs(args.output_dir)
    
    # Run evaluation
    eval_result = evaluate(
        data_path=args.data_path,
        output_dir=args.output_dir,
        start_idx=args.start_idx,
        end_idx=args.end_idx,
        verbose=args.verbose,
    )
    
    # Print summary
    summary = eval_result["summary"]
    print("\n" + "=" * 60)
    print("EVALUATION RESULTS")
    print("=" * 60)
    print(f"Total Instances:        {summary['total_instances']}")
    print(f"Fully Correct:          {summary['fully_correct_instances']}")
    print(f"Instance Accuracy:      {summary['instance_accuracy']*100:.1f}%")
    print(f"Total Test Cases:       {summary['total_test_cases']}")
    print(f"Passed Test Cases:      {summary['passed_test_cases']}")
    print(f"Test Case Accuracy:     {summary['test_case_accuracy']*100:.1f}%")
    print(f"Avg Soft Score:         {summary['avg_soft_score']*100:.1f}%")
    print(f"Avg Hard Score:         {summary['avg_hard_score']*100:.1f}%")
    print("=" * 60)
    
    # Save results
    results_file = args.results_file or os.path.join(args.output_dir, "eval_results.json")
    with open(results_file, "w") as f:
        json.dump(eval_result, f, indent=2)
    print(f"Results saved to: {results_file}")


if __name__ == "__main__":
    main()
